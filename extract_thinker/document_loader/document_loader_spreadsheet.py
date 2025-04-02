import io
from typing import Any, Dict, List, Union, Optional, Tuple
from io import BytesIO
from operator import attrgetter
from cachetools import cachedmethod
from cachetools.keys import hashkey
from extract_thinker.document_loader.cached_document_loader import CachedDocumentLoader


class DocumentLoaderSpreadSheet(CachedDocumentLoader):
    """Document loader for spreadsheet files."""
    
    SUPPORTED_FORMATS = ['xls', 'xlsx', 'xlsm', 'xlsb', 'odf', 'ods', 'odt', 'csv']

    def __init__(self, content=None, cache_ttl=300):
        """Initialize loader.
        
        Args:
            content: Initial content
            cache_ttl: Cache time-to-live in seconds
        """
        # Check required dependencies
        self._check_dependencies()
        super().__init__(content, cache_ttl)

    @staticmethod
    def _check_dependencies():
        """Check if required dependencies are installed."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Could not import openpyxl python package. "
                "Please install it with `pip install openpyxl xlrd`."
            )
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "Could not import xlrd python package. "
                "Please install it with `pip install openpyxl xlrd`."
            )

    def _get_openpyxl(self):
        """Lazy load openpyxl."""
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise ImportError(
                "Could not import openpyxl python package. "
                "Please install it with `pip install openpyxl`."
            )

    @cachedmethod(cache=attrgetter('cache'), 
                  key=lambda self, source: hashkey(source if isinstance(source, str) else source.getvalue(), self.vision_mode))
    def load(self, source: Union[str, BytesIO]) -> List[Dict[str, Any]]:
        """
        Load content from a spreadsheet and convert it to our standard format.
        Each sheet is treated as a separate "page" for consistency.

        Args:
            source: Either a file path or BytesIO stream
            
        Returns:
            List[Dict[str, Any]]: List of sheets, each containing content and sheet data
        """
        if not self.can_handle(source):
            raise ValueError(f"Cannot handle source: {source}")

        openpyxl = self._get_openpyxl()

        try:
            # Load workbook based on source type with data_only=True to get calculated values
            if isinstance(source, str):
                workbook = openpyxl.load_workbook(source, data_only=True)
            else:
                # BytesIO stream
                workbook = openpyxl.load_workbook(filename=BytesIO(source.read()), data_only=True)

            # Convert to our standard format
            sheets = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract plain text content with values instead of formulas
                sheet_content = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert all to strings
                    formatted_row = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell for cell in formatted_row):  # Skip empty rows
                        sheet_content.append(" | ".join(formatted_row))
                
                # Create a sheet entry
                sheet_dict = {
                    "content": "\n".join(sheet_content),
                    "image": None,
                    "name": sheet_name,
                    "is_spreadsheet": True  # Flag to indicate special handling
                }
                
                # # For backward compatibility, also keep the original data
                # row_data = [self._process_row(row) for row in sheet.iter_rows(values_only=True)]
                # sheet_dict["data"] = row_data
                
                sheets.append(sheet_dict)

            return sheets

        except Exception as e:
            raise ValueError(f"Error loading spreadsheet: {str(e)}")

    def _process_row(self, row: tuple) -> List[str]:
        """Process a row of spreadsheet data."""
        if all(cell in (None, '', ' ') for cell in row):
            return ["\n"]
        return [str(cell) if cell not in (None, '', ' ') else "" for cell in row]

    def can_handle_vision(self, source: Union[str, BytesIO]) -> bool:
        """Spreadsheet files don't support vision mode directly."""
        return False
        
    def convert_to_image(self, source: Union[str, BytesIO], 
                        max_width: int = 1200, 
                        max_height: int = 1800) -> Optional[Dict[str, bytes]]:
        """
        Convert spreadsheet to image(s) for visualization.
        
        Args:
            source: Spreadsheet file path or BytesIO object
            max_width: Maximum width of the rendered image
            max_height: Maximum height of the rendered image
            
        Returns:
            Dict mapping sheet names to PNG image bytes, or None if conversion fails
        """
        try:
            # Check for required libraries
            import pandas as pd
            from matplotlib import pyplot as plt
            from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
            
            # Load spreadsheet into pandas DataFrames
            if isinstance(source, str):
                xls = pd.ExcelFile(source)
            else:
                # Make sure we're at the beginning of the stream
                source.seek(0)
                xls = pd.ExcelFile(source)
            
            # Process each sheet
            sheet_images = {}
            for sheet_name in xls.sheet_names:
                # Read the sheet
                df = pd.read_excel(xls, sheet_name)
                
                # Create figure and axes
                fig_width, fig_height = self._calculate_figure_size(df, max_width, max_height)
                fig = plt.figure(figsize=(fig_width, fig_height), dpi=100)
                ax = plt.subplot()
                
                # Hide axes
                ax.axis('off')
                
                # Create table
                table = ax.table(
                    cellText=df.values,
                    colLabels=df.columns,
                    cellLoc='center',
                    loc='center'
                )
                
                # Style the table
                table.auto_set_font_size(False)
                table.set_fontsize(10)
                table.scale(1.2, 1.5)
                
                # Save to BytesIO
                buf = BytesIO()
                canvas = FigureCanvas(fig)
                canvas.print_png(buf)
                buf.seek(0)
                
                # Store the image
                sheet_images[sheet_name] = buf.getvalue()
                
                # Close the figure to free memory
                plt.close(fig)
            
            return sheet_images
            
        except ImportError as e:
            # Missing dependencies
            print(f"Warning: Could not convert spreadsheet to image. Missing dependency: {str(e)}")
            print("Please install with: pip install pandas matplotlib")
            return None
        except Exception as e:
            print(f"Error converting spreadsheet to image: {str(e)}")
            return None
            
    def _calculate_figure_size(self, df, max_width: int, max_height: int) -> Tuple[int, int]:
        """Calculate appropriate figure size based on DataFrame dimensions."""
        rows, cols = df.shape
        
        # Base sizes
        width = min(2 + (cols * 1.2), max_width/100)  # Adjust width based on columns
        height = min(2 + (rows * 0.5), max_height/100)  # Adjust height based on rows
        
        return width, height

    def convert_to_pdf(self, source: Union[str, BytesIO], 
                    output_path: str = None) -> Optional[BytesIO]:
        """
        Convert spreadsheet to PDF.
        
        Args:
            source: Spreadsheet file path or BytesIO object
            output_path: Path to save PDF output (optional)
            
        Returns:
            BytesIO object containing PDF data, or None if conversion fails
        """
        try:
            # Check for required libraries
            import pandas as pd
            from fpdf import FPDF
            
            # Load spreadsheet into pandas DataFrames
            if isinstance(source, str):
                xls = pd.ExcelFile(source)
            else:
                # Make sure we're at the beginning of the stream
                source.seek(0)
                xls = pd.ExcelFile(source)
                
            # Create PDF object
            pdf = FPDF()
            
            # Process each sheet
            for sheet_name in xls.sheet_names:
                # Read the sheet
                df = pd.read_excel(xls, sheet_name)
                
                # Add a page
                pdf.add_page()
                
                # Add sheet name as title
                pdf.set_font('Helvetica', 'B', 16)
                pdf.cell(0, 10, f'Sheet: {sheet_name}', 0, 1, 'C')
                pdf.ln(10)
                
                # Configure table
                pdf.set_font('Helvetica', 'B', 10)
                
                # Calculate column widths based on content
                col_widths = []
                for col_name in df.columns:
                    # Default minimum width
                    width = max(20, min(40, len(str(col_name)) * 4))
                    col_widths.append(width)
                
                # Table header
                for i, col_name in enumerate(df.columns):
                    pdf.cell(col_widths[i], 10, str(col_name), 1, 0, 'C')
                pdf.ln()
                
                # Table data
                pdf.set_font('Helvetica', '', 10)
                for _, row in df.iterrows():
                    for i, value in enumerate(row):
                        pdf.cell(col_widths[i], 10, str(value), 1, 0, 'C')
                    pdf.ln()
                    
            # Output PDF
            if output_path:
                pdf.output(output_path)
                
            # Return as BytesIO
            output = BytesIO()
            pdf.output(output)
            output.seek(0)
            return output
            
        except ImportError as e:
            # Missing dependencies
            print(f"Warning: Could not convert spreadsheet to PDF. Missing dependency: {str(e)}")
            print("Please install with: pip install pandas fpdf2")
            return None
        except Exception as e:
            print(f"Error converting spreadsheet to PDF: {str(e)}")
            return None
